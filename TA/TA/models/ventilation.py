from odoo import models, fields, api
from odoo.exceptions import ValidationError
import base64
import csv
import io
import chardet
import logging
import json

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

class Ventilation(models.Model):
    _name = 'ventilation.model'
    _description = 'Modèle de Ventilation'

    dossier_id = fields.Many2one('ta.dossier', string='Dossier lié')

    # Champs related pour afficher les informations du dossier
    sequence_dossier = fields.Char(related='dossier_id.sequence', string='Num Dossier', readonly=True)
    state_dossier = fields.Selection(related='dossier_id.state', string='Statut Dossier', readonly=True)
    date_dossier = fields.Date(related='dossier_id.date', string='Date Dossier', readonly=True)
    loading_port_id = fields.Selection(related='dossier_id.loading_port_id', string='Port de Chargement', readonly=True)
    shipper_id = fields.Many2one(related='dossier_id.shipper_id', string='Shipper du Dossier', readonly=True)
    consignee_id = fields.Many2one(related='dossier_id.consignee_id', string='Consignee du Dossier', readonly=True)
    poids_brut = fields.Float(related='dossier_id.poids_brut', string='Poids Brut du Dossier (kg)', readonly=True)
    nombre = fields.Char(related='dossier_id.nombre', string='Colisage du Dossier', readonly=True)

    # Champs de base
    file_data = fields.Binary(string='Importer CSV', help="Téléchargez ici le fichier CSV")
    file_name = fields.Char(string='Nom du fichier')
    val_coeff = fields.Float(string='Coefficient', compute='_val_coeff', store=True)
    valeur_devise = fields.Float(string='VALEUR DEVISE')
    base_coef = fields.Float(string='BASE COEF')
    devise = fields.Selection([
        ('usd', 'USD - Dollar américain'),
        ('eur', 'EUR - Euro'),
        ('gbp', 'GBP - Livre sterling'),
        ('jpy', 'JPY - Yen japonais'),
        ('cad', 'CAD - Dollar canadien'),
        ('aud', 'AUD - Dollar australien'),
        ('chf', 'CHF - Franc suisse'),
        ('cny', 'CNY - Yuan chinois'),
        ('inr', 'INR - Roupie indienne'),
        ('sgd', 'SGD - Dollar de Singapour'),
        ('nzd', 'NZD - Dollar néo-zélandais'),
        ('sek', 'SEK - Couronne suédoise'),
        ('dkk', 'DKK - Couronne danoise'),
        ('nok', 'NOK - Couronne norvégienne'),
        ('krw', 'KRW - Won sud-coréen'),
        ('rub', 'RUB - Rouble russe'),
        ('brl', 'BRL - Réal brésilien'),
        ('zar', 'ZAR - Rand sud-africain'),
        ('try', 'TRY - Lire turque'),
        ('mxn', 'MXN - Peso mexicain'),
    ], string='DEVISE')
    fret = fields.Float(string='Fret')
    aconage = fields.Float(string='Aconage')
    vent_incoterm = fields.Selection([
        ('exw', 'EXW - Ex Works (lieu convenu)'),
        ('fca', 'FCA - Free Carrier (lieu convenu)'),
        ('cpt', 'CPT - Carriage Paid To (lieu de destination convenu)'),
        ('cip', 'CIP - Carriage and Insurance Paid To (lieu de destination convenu)'),
        ('dap', 'DAP - Delivered At Place (lieu de destination convenu)'),
        ('dpu', 'DPU - Delivered at Place Unloaded (lieu de destination convenu)'),
        ('ddp', 'DDP - Delivered Duty Paid (lieu de destination convenu)'),
        ('fob', 'FOB - Free On Board (port d’embarquement convenu)'),
        ('cfr', 'CFR - Cost and Freight (port de destination convenu)'),
        ('cif', 'CIF - Cost, Insurance and Freight (port de destination convenu)'),
    ], string='Incoterm')
    taux_de_change = fields.Float(string='Taux de change', store=True)

    # Champs calculés
    pivot_data = fields.Text(string='Données Pivot', compute='_compute_pivot_data', store=True)
    pivot_table = fields.Html(string='Tableau Pivot', compute='_compute_pivot_data', store=True)
    cv = fields.Float(string='CV (Contrevaleur)', compute='_compute_cv', store=True)
    assurance = fields.Float(string="Assurance", compute="_compute_assurance")
    valeur_a_declarer = fields.Float(string='Valeur à Déclarer', compute='_compute_valeur_a_declarer')

    @api.constrains('taux_de_change', 'valeur_devise')
    def _check_taux_de_change(self):
        for record in self:
            if record.taux_de_change <= 0 or record.valeur_devise < 0:
                raise ValidationError("Taux de change et valeur devise doivent être positifs.")

    @api.depends('valeur_devise', 'devise', 'fret', 'vent_incoterm')
    def _compute_cv(self):
        for record in self:
            vent_incoterm = str(record.vent_incoterm).lower()  # Convertir en minuscules
            if vent_incoterm in ('exw', 'fca', 'fas', 'fob'):
                record.cv = record.valeur_devise * record.taux_de_change + record.fret
            else:
                record.cv = record.valeur_devise * record.taux_de_change

    @api.depends('cv', 'assurance', 'aconage')
    def _compute_valeur_a_declarer(self):
        for record in self:
            record.valeur_a_declarer = record.cv + record.assurance + record.aconage

    @api.depends('cv')
    def _compute_assurance(self):
        for record in self:
            record.assurance = record.cv * 0.0003

    @api.depends('base_coef', 'valeur_a_declarer')
    def _val_coeff(self):
        for record in self:
            if record.base_coef != 0:  # Vérifiez que base_coef n'est pas égal à zéro
                record.val_coeff = record.valeur_a_declarer / record.base_coef
            else:
                record.val_coeff = 0

    @api.depends('lignes_ventilation')
    def _compute_pivot_data(self):
        for record in self:
            pivot_data = {}
            for ligne in record.lignes_ventilation:
                # Clé unique pour chaque combinaison de valeurs
                key = f"{ligne.hscode}-{ligne.accord}-{ligne.origine}-{ligne.unite}-{ligne.ap_sp}"
                if key not in pivot_data:
                    pivot_data[key] = {
                        'hscode': ligne.hscode,
                        'accord': ligne.accord,
                        'origine': ligne.origine,
                        'unite': ligne.unite,
                        'ap_sp': ligne.ap_sp,
                        'quantite_total': 0.0,
                        'valeur_devise_total': 0.0,
                        'poids_net_total': 0.0,
                    }
                pivot_data[key]['quantite_total'] += ligne.quantite
                pivot_data[key]['valeur_devise_total'] += ligne.valeur_devise
                pivot_data[key]['poids_net_total'] += ligne.poids_net

            # Sérialisation en JSON
            record.pivot_data = json.dumps(pivot_data)

            # Convertir pivot_data en dictionnaire
            pivot_data = json.loads(record.pivot_data) if record.pivot_data else {}

            # Génération du tableau HTML
            table_html = '<table class="table table-condensed"><thead><tr>'
            table_html += '<th>Hscode</th><th>Accord</th><th>Origine</th><th>Unité</th><th>AP/SP</th>'
            table_html += '<th>Quantité Total</th><th>Valeur Devise Total</th><th>Poids Net Total</th>'
            table_html += '</tr></thead><tbody>'
            for data in pivot_data.values():
                table_html += '<tr>'
                table_html += f'<td>{data["hscode"]}</td><td>{data["accord"]}</td><td>{data["origine"]}</td>'
                table_html += f'<td>{data["unite"]}</td><td>{data["ap_sp"]}</td>'
                table_html += f'<td>{data["quantite_total"]}</td><td>{data["valeur_devise_total"]}</td><td>{data["poids_net_total"]}</td>'
                table_html += '</tr>'
            table_html += '</tbody></table>'

            record.pivot_table = table_html

    lignes_ventilation = fields.One2many('ventilation.ligne', 'ventilation_id', string='Lignes de Ventilation')
    def import_file(self):
        if not self.file_data or not self.file_name:
            raise ValidationError("Aucun fichier n'a été téléchargé ou le nom de fichier est invalide.")

        file_content = base64.b64decode(self.file_data)

        VentilationLigne = self.env['ventilation.ligne']
        if self.file_name.endswith('.csv'):
            # Utiliser chardet pour détecter l'encodage du fichier CSV
            result = chardet.detect(file_content)
            file_encoding = result['encoding'] or 'utf-8'

            try:
                decoded_content = file_content.decode(file_encoding)
            except UnicodeDecodeError:
                raise ValidationError(f"Impossible de décoder le fichier avec l'encodage {file_encoding}.")

            file_io = io.StringIO(decoded_content)
            csv_reader = csv.DictReader(file_io, delimiter=',')
            for row in csv_reader:
                # Création d'un enregistrement pour chaque ligne du CSV
                VentilationLigne.create({
                    'hscode': row.get('hscode'),
                    'description': row.get('description'),
                    'accord': row.get('accord'),
                    'origine': row.get('origine'),
                    'unite': row.get('unite'),
                    'quantite': float(row.get('quantite', 0)),
                    'valeur_devise': float(row.get('valeur_devise', 0)),
                    'poids_net': float(row.get('poids_net', 0)),
                    'ventilation_id': self.id,
                })
        elif self.file_name.endswith('.xls'):
            if not xlrd:
                raise ValidationError("La bibliothèque xlrd n'est pas installée.")
            workbook = xlrd.open_workbook(file_contents=file_content)
            sheet = workbook.sheet_by_index(0)
            headers = sheet.row_values(0)
            for row_idx in range(1, sheet.nrows):
                row = sheet.row_values(row_idx)
                data = dict(zip(headers, row))
                VentilationLigne.create({
                    'hscode': data.get('hscode'),
                    'description': data.get('description'),
                    'accord': data.get('accord'),
                    'origine': data.get('origine'),
                    'unite': data.get('unite'),
                    'quantite': float(data.get('quantite', 0)),
                    'valeur_devise': float(data.get('valeur_devise', 0)),
                    'poids_net': float(data.get('poids_net', 0)),
                    'ventilation_id': self.id,
                })
        elif self.file_name.endswith('.xlsx'):
            if not openpyxl:
                raise ValidationError("La bibliothèque openpyxl n'est pas installée.")
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for row in sheet.iter_rows(min_row=2):
                data = {headers[i]: cell.value for i, cell in enumerate(row)}
                VentilationLigne.create({
                    'hscode': data.get('hscode'),
                    'description': data.get('description'),
                    'accord': data.get('accord'),
                    'origine': data.get('origine'),
                    'unite': data.get('unite'),
                    'quantite': float(data.get('quantite', 0)),
                    'valeur_devise': float(data.get('valeur_devise', 0)),
                    'poids_net': float(data.get('poids_net', 0)),
                    'ventilation_id': self.id,
                })

        else:
            raise ValidationError("Format de fichier non pris en charge.")

        # Effacer le champ du fichier après l'importation
        self.file_data = False
        self.file_name = False

class VentilationLigne(models.Model):
    _name = 'ventilation.ligne'
    _description = 'Ligne de Ventilation'

    # Champs de la ligne de ventilation
    hscode = fields.Char(string='Hscode', size=10)
    description = fields.Char(string='Description', size=30)
    accord = fields.Selection(selection=[('usa', 'USA'), ('eu', 'UE'), ('aele', 'AELE'), ('ae', 'AE')], string='Accord')
    origine = fields.Selection([
        ('france', 'France'),
        ('espagne', 'Espagne'),
        ('italie', 'Italie'),
        ('allemagne', 'Allemagne'),
        ('royaume-uni', 'Royaume-Uni'),
        ('pays-bas', 'Pays-Bas'),
        ('belgique', 'Belgique'),
        ('portugal', 'Portugal'),
        ('suisse', 'Suisse'),
        ('tunisie', 'Tunisie'),
        ('algerie', 'Algérie'),
        ('egypte', 'Égypte'),
        ('usa', 'États-Unis'),
        ('canada', 'Canada'),
        ('chine', 'Chine'),
        ('japon', 'Japon'),
        ('inde', 'Inde'),
        ('bresil', 'Brésil'),
        ('argentine', 'Argentine'),
        ('afrique-du-sud', 'Afrique du Sud'),
        ('nigeria', 'Nigeria'),
        ('kenya', 'Kenya'),
        ('australie', 'Australie'),
        ('nouvelle-zelande', 'Nouvelle-Zélande'),
        ('singapour', 'Singapour'),
        ('malaisie', 'Malaisie'),
        ('emirats-arabes-unis', 'Émirats arabes unis'),
        ('qatar', 'Qatar'),
        ('arabie-saoudite', 'Arabie saoudite'),
        ('iran', 'Iran'),
        ('turquie', 'Turquie'),
        ('russie', 'Russie'),
        ('ukraine', 'Ukraine'),
        ('pologne', 'Pologne'),
        ('republique-tcheque', 'République tchèque'),
        ('slovaquie', 'Slovaquie'),
        ('hongrie', 'Hongrie'),
        ('autriche', 'Autriche'),
        ('grece', 'Grèce'),
        ('suede', 'Suède'),
        ('norvege', 'Norvège'),
        ('finlande', 'Finlande'),
        ('danemark', 'Danemark'),
        ('portugal', 'Portugal'),
        ('maroc', 'Maroc'),
    ], string='Origine')
    unite = fields.Selection(selection=[
        ('kg', 'Kilogramme (kg)'),
        ('g', 'Gramme (g)'),
        ('mg', 'Milligramme (mg)'),
        ('tonne', 'Tonne (tonne)'),
        ('litre', 'Litre (litre)'),
        ('ml', 'Millilitre (ml)'),
        ('m3', 'Mètre cube (m³)'),
        ('m2', 'Mètre carré (m²)'),
        ('m', 'Mètre (m)'),
        ('cm', 'Centimètre (cm)'),
        ('mm', 'Millimètre (mm)'),
        ('km', 'Kilomètre (km)'),
        ('ft', 'Pied (ft)'),
    ], string='Unité')
    quantite = fields.Float(string='Quantité')
    valeur_devise = fields.Float(string='Valeur article devise')
    poids_net = fields.Float(string='Poids net')
    valeur_mad = fields.Float(string='Valeur article MAD', compute='_compute_valeur_mad')
    ap_sp = fields.Selection(
        [('AP', 'AP'), ('SP', 'SP')],
        string='AP/SP',
        default='AP',
    )

    # Relation avec le modèle Ventilation
    ventilation_id = fields.Many2one('ventilation.model', string='Ventilation')

    @api.constrains('valeur_devise')
    def _check_valeur_devise(self):
        for record in self:
            if record.valeur_devise < 0:
                raise ValidationError("Valeur devise doit être positive.")

    @api.depends('ventilation_id.val_coeff', 'valeur_devise')
    def _compute_valeur_mad(self):
        for record in self:
            record.valeur_mad = record.ventilation_id.val_coeff * record.valeur_devise